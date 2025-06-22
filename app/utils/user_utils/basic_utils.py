import app.keyboards.user_keyboards as kb
import time

from openpyxl import load_workbook

from CONSTANTS import *


def get_data(filename):
    filename = filename.split('_')
    filename.pop(0)
    filename = '-'.join(filename)

    book = load_workbook(filename=f"{PATH_FOLDER}/{filename}-sm.xlsx")

    sheet = book.active
    max_row = sheet.max_row


    sheet = book["1"]

    menu = {}

    for i in range(4, max_row):
        line = sheet["D" + str(i)]

        if line.value:
            dish = line.value
            category = sheet["B" + str(i)].value

            menu[category] = dish

    return menu


async def change_message(bot, message_chat_id, message_id):
    print("work")
    result = await kb.main()

    if result[1].export():
        print(result)
        result = await kb.main()

        await bot.edit_message_text(text=TEXT, chat_id=message_chat_id, message_id=message_id+2, reply_markup=result[0])

    else:
        await bot.edit_message_text(text=TEXT_VACATION, chat_id=message_chat_id, message_id=message_id+2)
